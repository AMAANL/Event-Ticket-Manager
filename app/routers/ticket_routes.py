import csv
import json
import uuid
from datetime import date, datetime
from io import BytesIO, StringIO

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from .. import models
from ..deps import get_current_user, get_db
from ..qr_utils import generate_qr_bytes, sign


router = APIRouter()


def _json_safe(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value is None:
        return ""
    return value


def _clean_header(value) -> str:
    return str(value or "").strip()


def _normalize_rows(rows):
    normalized = []
    fields = []

    for row in rows:
        cleaned = {}
        for key, value in row.items():
            clean_key = _clean_header(key)
            if not clean_key:
                continue
            if clean_key not in fields:
                fields.append(clean_key)
            cleaned[clean_key] = _json_safe(value)

        if any(str(value).strip() for value in cleaned.values()):
            normalized.append(cleaned)

    return fields, normalized


def _parse_csv(content: bytes):
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(StringIO(text))
    return _normalize_rows(reader)


def _parse_json(content: bytes):
    payload = json.loads(content.decode("utf-8-sig"))
    if isinstance(payload, dict):
        payload = payload.get("attendees") or payload.get("rows") or payload.get("data")

    if not isinstance(payload, list) or not all(isinstance(row, dict) for row in payload):
        raise ValueError("JSON must be a list of attendee objects, or an object with attendees/rows/data.")

    return _normalize_rows(payload)


def _parse_xlsx(content: bytes):
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)

    try:
        headers = [_clean_header(value) for value in next(rows)]
    except StopIteration as exc:
        raise ValueError("Spreadsheet is empty.") from exc

    mapped_rows = []
    for row in rows:
        mapped_rows.append(dict(zip(headers, row)))

    return _normalize_rows(mapped_rows)


def _parse_attendee_file(file_name: str, content: bytes):
    lower_name = file_name.lower()
    if lower_name.endswith(".csv"):
        return _parse_csv(content)
    if lower_name.endswith(".json"):
        return _parse_json(content)
    if lower_name.endswith(".xlsx"):
        return _parse_xlsx(content)

    raise ValueError("Upload a CSV, JSON, or XLSX file.")


def _event_for_user(event_id: int, user_id: int, db: Session):
    return (
        db.query(models.Event)
        .filter_by(id=event_id, organizer_id=user_id)
        .first()
    )


@router.post("/ticket/{event_id}")
def issue_ticket(
    event_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    event = _event_for_user(event_id, user_id, db)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    ticket_id = str(uuid.uuid4())
    ticket = models.Ticket(ticket_id=ticket_id, event_id=event_id, attendee_data={})
    db.add(ticket)
    db.commit()

    _, payload = generate_qr_bytes(ticket_id, event_id)

    return {
        "ticket_id": ticket_id,
        "qr_url": f"/tickets/{ticket_id}/qr",
        "qr_json": payload,
        "attendee_data": ticket.attendee_data,
    }


@router.post("/event/{event_id}/import")
async def import_attendees(
    event_id: int,
    file: UploadFile = File(...),
    replace_existing: bool = Form(False),
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    event = _event_for_user(event_id, user_id, db)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    try:
        fields, rows = _parse_attendee_file(file.filename or "", await file.read())
    except Exception as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc

    if not rows:
        raise HTTPException(status_code=400, detail="No attendee rows found.")

    if replace_existing:
        old_tickets = db.query(models.Ticket).filter_by(event_id=event_id).all()
        for ticket in old_tickets:
            db.delete(ticket)

    event.field_schema = fields
    tickets = []

    for attendee_data in rows:
        ticket_id = str(uuid.uuid4())
        ticket = models.Ticket(
            ticket_id=ticket_id,
            event_id=event_id,
            attendee_data=attendee_data,
        )
        db.add(ticket)
        _, payload = generate_qr_bytes(ticket_id, event_id)
        tickets.append(
            {
                "ticket_id": ticket_id,
                "qr_url": f"/tickets/{ticket_id}/qr",
                "qr_json": payload,
                "attendee_data": attendee_data,
            }
        )

    db.commit()

    return {
        "event_id": event_id,
        "imported": len(tickets),
        "fields": fields,
        "sample_tickets": tickets[:25],
        "sample_limit": 25,
    }


@router.get("/event/{event_id}/tickets")
def list_tickets(
    event_id: int,
    db: Session = Depends(get_db),
    user_id: int = Depends(get_current_user),
):
    event = _event_for_user(event_id, user_id, db)
    if not event:
        raise HTTPException(status_code=404, detail="Event not found")

    tickets = (
        db.query(models.Ticket)
        .filter_by(event_id=event_id)
        .order_by(models.Ticket.id.desc())
        .all()
    )

    return [
        {
            "ticket_id": ticket.ticket_id,
            "event_id": ticket.event_id,
            "used": ticket.used,
            "attendee_data": ticket.attendee_data,
            "qr_url": f"/tickets/{ticket.ticket_id}/qr",
            "qr_json": {
                "ticket_id": ticket.ticket_id,
                "event_id": ticket.event_id,
                "signature": sign(ticket.ticket_id),
            },
        }
        for ticket in tickets
    ]


@router.get("/tickets/{ticket_id}/qr", include_in_schema=False)
def get_ticket_qr(ticket_id: str, db: Session = Depends(get_db)):
    """Generate and serve QR code for a ticket on-demand"""
    ticket = db.query(models.Ticket).filter(
        models.Ticket.ticket_id == ticket_id
    ).first()

    if not ticket:
        raise HTTPException(status_code=404, detail="Ticket not found")

    qr_bytes, _ = generate_qr_bytes(ticket.ticket_id, ticket.event_id)

    return StreamingResponse(
        BytesIO(qr_bytes),
        media_type="image/png",
        headers={"Content-Disposition": f"inline; filename={ticket_id}.png"}
    )
